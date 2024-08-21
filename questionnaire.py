import json
import csv
import openpyxl

# 加载JSON数据
with open('lens.json', 'r') as file:
    data = json.load(file)

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 创建CSV文件并获取写入对象
csv_file = open('workload.csv', 'w', newline='', encoding='utf-8')
csv_writer = csv.writer(csv_file)

# 写入标题行
# headers = ['Pillar ID', 'Question ID', 'Choice ID', 'Pillar Name', 'Question Title', 'Choice Title', 'Question Description', 'helpfulResource', 'WA_checked', 'resource_comments']
headers = ['Pillar Name', 'Question Title', 'Choice Title', 'Question ID', 'Choice ID', 'WA_checked', 'resource_comments']
csv_writer.writerow(headers)

# 遍历pillars、questions和choices,并写入CSV文件和Excel工作表
for pillar in data['pillars']:
    pillar_id = pillar['id']
    pillar_name = pillar['name']

    for question in pillar['questions']:
        question_id = question['id']
        question_title = question['title']
        question_description = question['description']

        for choice in question['choices']:
            choice_id = choice['id']
            choice_title = choice['title']

            # 获取 helpfulResource 字段的值
            helpful_resource = choice.get('helpfulResource', '')
            helpful_resource_str = str(helpful_resource)  # 将 helpful_resource 转换为字符串

            # 检查 helpfulResource 是否包含 "Trusted Advisor Checks"
            wa_checked = "Yes" if "Trusted Advisor Checks" in helpful_resource_str else "No"

            # 获取 resourceComments 字段的值
            resource_comments = ''
            if helpful_resource and isinstance(helpful_resource, dict) and 'displayText' in helpful_resource:
                resource_comments = helpful_resource['displayText']

            # row = [pillar_id, question_id, choice_id, pillar_name, question_title, choice_title, question_description, helpful_resource_str, wa_checked, resource_comments]
            row = [pillar_name, question_title, choice_title, question_id, choice_id, wa_checked, resource_comments]
            csv_writer.writerow(row)

            # 将数据写入对应的 Excel 工作表
            sheet_name = pillar_name
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
                sheet = workbook[sheet_name]
                sheet.append(headers)  # 在新工作表中添加标题行
            sheet = workbook[sheet_name]
            sheet.append(row)

# 关闭CSV文件
csv_file.close()

# 保存 Excel 工作簿
workbook.save('workload.xlsx')
